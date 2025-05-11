import tkinter as tk
from tkinter import messagebox
from tkinter.simpledialog import askstring
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import re
import os
import time



def extract_amazon_price(driver):
    try:
        # This selector gets the full price (including dollars and cents)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "span.a-offscreen"))
        )
        elements = driver.find_elements(By.CSS_SELECTOR, "span.a-offscreen")
        
        for elem in elements:
            raw_price = elem.text.strip().replace("$", "").replace(",", "")
            try:
                # Try converting to float to validate format
                float(raw_price)
                return raw_price
            except ValueError:
                continue
    except TimeoutException:
        pass
    except Exception as e:
        print(f"Error extracting price: {e}")
    
    return "0.00"

def scrape_and_export():
    event_name = askstring("Event Name", "What event is this for?")
    if not event_name or not event_name.strip():
        messagebox.showerror("Error", "Event name is required.")
        return

    safe_event_name = re.sub(r'[^A-Za-z0-9_]', '_', event_name.strip())
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    filename = os.path.join(desktop_path, f"{safe_event_name}_invoice.xlsx")

    urls = text_input.get("1.0", tk.END).strip().split("\n")

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")

    try:
        driver = webdriver.Chrome(options=options)
    except WebDriverException as e:
        messagebox.showerror("Error", f"ChromeDriver failed to start:\n{e}")
        return

    data = []

    for url in urls:
        url = url.strip()
        if not url or "amazon" not in url:
            continue

        try:
            driver.get(url)
            time.sleep(5)

            with open("debug_page.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            driver.save_screenshot("screenshot.png")

            try:
                title_elem = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "productTitle"))
                )
                title = title_elem.text.strip()
            except:
                title = "Title not found"

            price = extract_amazon_price(driver)
            data.append([title, price, "", "", url])
        except Exception as e:
            data.append([f"Error: {str(e)}", "0.00", "", "", url])

    driver.quit()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Scraped Data"

    # Headers
    header = ["Item", "Quantity", "Price", "Total Price", "Comments", "URL"]
    ws.append(header)
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, size=15)

    # Data 
    for idx, row in enumerate(data, start=2):
        title, price, _, _, url = row

        # Item with hyperlink
        item_cell = ws.cell(row=idx, column=1)
        item_cell.value = title
        item_cell.hyperlink = url
        item_cell.style = "Hyperlink"

        # Quantity
        ws.cell(row=idx, column=2).value = ""

        # Adjusted Price
        try:
            adjusted_price = round(float(price), 2)
        except ValueError:
            adjusted_price = 1.00
        price_cell = ws.cell(row=idx, column=3)
        price_cell.value = adjusted_price
        price_cell.number_format = '"$"#,##0.00'

        # Total Price
        total_cell = ws.cell(row=idx, column=4)
        total_cell.value = f"=B{idx}*C{idx}"
        total_cell.number_format = '"$"#,##0.00'

        # Comments (empty)
        ws.cell(row=idx, column=5).value = ""

        # URL
        ws.cell(row=idx, column=6).value = url

    # Final total row
    total_row = len(data) + 2
    ws[f"B{total_row}"] = "Total Cost:"
    total_cell = ws[f"D{total_row}"]
    total_cell.value = f"=SUM(D2:D{total_row - 1})"
    total_cell.number_format = '"$"#,##0.00'

    # Set column width for item 
    ws.column_dimensions["A"].width = 40 
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="center")

    #Set widths for other columns
    ws.column_dimensions["B"].width = 10  # Quantity
    ws.column_dimensions["C"].width = 12  # Price
    ws.column_dimensions["D"].width = 15  # Total Price
    ws.column_dimensions["E"].width = 20  # Comments
    ws.column_dimensions["F"].width = 60  # URL

    # Text Vertical aline center
    for col in range(1, 7): 
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                cell.alignment = cell.alignment.copy(wrap_text=True, vertical="center")

    try:
        wb.save(filename)
        messagebox.showinfo("Done", f"Data exported to {filename}")
        os.startfile(filename)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save file:\n{e}")

# GUI Setup
root = tk.Tk()
root.title("Prize Fetch")
root.geometry("600x400")

tk.Label(root, text="Enter Amazon Product URLs (one per line):").pack(pady=5)
text_input = tk.Text(root, height=15)
text_input.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

tk.Button(root, text="Scrape and Export", command=scrape_and_export).pack(pady=10)

root.mainloop()